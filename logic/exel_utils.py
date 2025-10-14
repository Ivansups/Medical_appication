import math
import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

DEFAULT_FILENAME = "patients.xlsx"


def calculate_ckd_epi(age, gender, creatinine):
    """Расчет СКФ по формуле CKD-EPI"""
    try:
        # Конвертируем креатинин из мкмоль/л в мг/дл (1 мг/дл = 88.4 мкмоль/л)
        scr_mg_dl = float(creatinine) / 88.4

        if gender == "Муж":
            k = 0.9
            alpha = -0.302
            gender_factor = 1.0
        else:
            k = 0.7
            alpha = -0.241
            gender_factor = 1.012

        scr_k = scr_mg_dl / k

        min_val = min(scr_k, 1)
        max_val = max(scr_k, 1)

        gfr = 142 * math.pow(min_val, alpha) * math.pow(max_val, -1.2) * math.pow(0.9938, age) * gender_factor
        return round(gfr)
    except (ValueError, TypeError, ZeroDivisionError):
        return None


def calculate_creatinine_clearance(age, weight, gender, creatinine):
    """Расчет клиренса креатинина по Кокрофту-Голту"""
    try:
        age = float(age)
        weight = float(weight)
        creatinine = float(creatinine)

        # Формула Кокрофта-Голта
        if gender == "Муж":
            ccr = ((140 - age) * weight) / (72 * creatinine / 88.4)
        else:
            ccr = ((140 - age) * weight) / (72 * creatinine / 88.4) * 0.85

        return round(ccr)
    except (ValueError, TypeError, ZeroDivisionError):
        return None


def create_or_load_workbook(filename=DEFAULT_FILENAME):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "Пол",
                "Возраст",
                "Вес",
                "Рост",
                "Креатинин",
                "Клиренс креатинина",
                "MPV",
                "PLCR",
                "Спонтанная агрегация",
                "Индуц. агрегация 1 мкМоль АДФ",
                "Индуц. агрегация 5 мкМоль АДФ",
                "Индуц. агрегация 15 мкл арахидоновой кислоты",
                "Генотип CYP2C19",
                "Генотип ABCB1",
                "Препараты",
                "Состояние агрегации",
                "Скорость выведения клопидогрела (ABCB1)",
                "Модуль 1",
                "Модуль 2",
                "Модуль 3",
                "Коэффициент прогноза",
                "Оценка прогноза",
            ]
        )
    return wb, ws


def autofit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column  # номер колонки (1, 2, 3...)
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except (AttributeError, TypeError):
                pass
        adjusted_width = max_length + 2  # небольшой запас
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width


def append_patient_data(filename, data_row):
    wb, ws = create_or_load_workbook(filename)
    if ws is not None:
        ws.append(data_row)
        autofit_columns(ws)
        wb.save(filename)
    else:
        raise ValueError("Не удалось создать или получить рабочий лист Excel")
